# This is a script to assist in a variety of tasks related to Cisco DNAC deploments
# === DESCRIPTION ===
# For any questions, please contact George Bekmezian, george.bekmezian@cvetech.com
#
# This file was last modified on 20211207

import requests
import csv
import logging
import os
import sys
import datetime
import inspect
import json
import os.path as path
import pprint
import ipaddress
import traceback
from logging.handlers import RotatingFileHandler
from datetime import timedelta
from requests.auth import HTTPBasicAuth
from getpass import getpass
from time import sleep
from dnacentersdk import DNACenterAPI, ApiError

#The next two lines are not best practice and are for lab use only
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

logfile = ''
#imported from login.py or requested from user via CLI
HOST = ''
password = ''
username = ''
#local pool site id
siteId = ''
#global pool ID
parentUuid = ''

def CheckUsage(argvlocal):
    Action = None
    if len(argvlocal) < 2:
        printusage(argvlocal)
        return Action
    Action = argvlocal[1]
    mylogger('%s this is the action %s\n' % (lineno(), Action))
    if Action.lower() == 'never happens':
        if len(argvlocal) != 4:
            printusage(argvlocal)
    elif Action.lower() == 'importpoolsfromcsv':
        if len(argvlocal) != 3:
            printusage(argvlocal)
            exit(1)
        elif not path.exists(argvlocal[2]):
            printusertext('%s csv file not found: %s' % (lineno(), argvlocal[2]))
            printusage(argvlocal)
            exit(1)
    elif Action.lower() == 'createglobalpool':
        pass
    elif Action.lower() == 'createdevicecredentials':
        pass
    elif Action.lower() == 'createnetworksettings':
        pass
    elif Action.lower() == 'importfromexcel':
        pass
    elif Action.lower() == 'deletefromexcel':
        pass
    else:
        printusage(argvlocal)
        printusertext("Didn't match any actions available actions")
    return Action

def printusage(argvlocal):
    #prints help text
    printusertext('Argument length: %s' % str(len(argvlocal)))
    printusertext('This is a script to perform repetitive DNAC tasks.')
    printusertext('Usage:')
    printusertext(f'python3 {argvlocal[0]} ImportPoolsFromCSV <CSV filename>')
    printusertext(f'python3 {argvlocal[0]} CreateGlobalPool')
    printusertext(f'python3 {argvlocal[0]} CreateDeviceCredentials')
    printusertext(f'python3 {argvlocal[0]} CreateNetworkSettings')
    printusertext(f'python3 {argvlocal[0]} ImportFromExcel')
    printusertext(f'python3 {argvlocal[0]} DeleteFromExcel')

def printusertext(p_message):
    # prints a line of text that is meant for the user to read
    # do not process these lines when chaining scripts
    print('@ %s' % p_message)

def initlogging(argvlocal):
    global logfile 
    mynow = datetime.datetime.now()
    mytimestamp = mynow.strftime("%Y%m%d_%H%M")
    logger = logging.getLogger('mylog')
    myfmt = logging.Formatter("%(asctime)-15s @ [%(threadName)s] @ %(message)s  ")
    if '/' in argvlocal[0]:
        print(argvlocal[0])
        myfilename = argvlocal[0][argvlocal[0].rfind('/')+1:len(argvlocal[0])]
    else:
        myfilename = argvlocal[0]
    myhandler = RotatingFileHandler('logs/%s_%s.log' % (mytimestamp, myfilename[0:myfilename.find('.',2)]), maxBytes=100000,
                                  backupCount=10)
    myhandler.setFormatter(myfmt)
    logger.setLevel(logging.INFO)
    logger.addHandler(myhandler)
    #this is to print it on exit
    logfile = myhandler.baseFilename

def mylogger(mymsg):
    logger = logging.getLogger('mylog')
    logger.info(mymsg)

def lineno():
    """Returns the current line number in our program."""
    return inspect.currentframe().f_back.f_lineno

def getAuthToken(username, password):
    HEADERS = {
        'content-type': "application/json"
    }

    URL = f'https://{HOST}/dna/system/api/v1/auth/token'
    response = requests.post(url=URL, headers=HEADERS, auth=HTTPBasicAuth(username, password), verify=False)
    TOKEN = response.json().get('Token')
    return TOKEN

def getSitesSDK(DNAC_SDK):
    #Returns site name hierarchy and site id
    sites = {}
    mylogger('About to retrieve list of sites')
    response = DNAC_SDK.sites.get_site()
    mylogger(f'response: {response}')
    for site in response.response:
        sites[site.get('siteNameHierarchy')] = site.get('id')
    return sites

def CreateSiteSDK(DNAC_SDK,*ARGS):
    PAYLOAD, SiteType = ARGS
    mylogger(f'About to create site with payload: {PAYLOAD}')
    response = DNAC_SDK.sites.create_site(PAYLOAD, SiteType)
    mylogger(f'response: {response}')
    return {"executionId": response.get('executionId')}

#deprecated. Use getSitesSDK
def getSiteId(SiteName, AuthKey):
    URL = f'https://{HOST}/dna/intent/api/v1/site'
    HEADERS = {
            'content-type': "application/json",
            'x-auth-token': AuthKey
        }
    response = requests.get(url=URL, headers=HEADERS, verify=False)
    mylogger(response.content)

    for item in response.json()['response']:
        if item['name'] == SiteName:
            mylogger(f'found site. id is: {item["id"]}')
            siteId = item["id"]
            return siteId
    mylogger(f'Failed to find site named: {SiteName}')
    return None

def getGlobalPoolId(GlobalPoolName, AuthKey):
    URL = f'https://{HOST}/dna/intent/api/v1/global-pool'
    HEADERS = {
            'content-type': "application/json",
            'x-auth-token': AuthKey
        }
    response = requests.get(url=URL, headers=HEADERS, verify=False)

    for item in response.json()['response']:
        if item["ipPoolName"] == GlobalPoolName:
            mylogger(f'found global pool. id is: {item["id"]}')
            parentUuid = item["id"]
            return parentUuid
    mylogger(f'Failed to find global pool named: {GlobalPoolName}')
    return None

def CreateGlobalPool(PAYLOAD,AuthKey):
    URL = f'https://{HOST}/dna/intent/api/v1/global-pool'
    HEADERS = {
            'content-type': "application/json",
            'x-auth-token': AuthKey
        }
    mylogger(f'About to create global pool with payload: {PAYLOAD}')
    response = requests.post(url=URL, headers=HEADERS, verify=False, data=json.dumps(PAYLOAD))
    mylogger(f'response.text: {response.text}')

def CreateGlobalPoolSDK(PAYLOAD,DNAC_SDK):
    mylogger(f'About to create global pool with payload: {PAYLOAD}')
    response = DNAC_SDK.network_settings.create_global_pool(payload=PAYLOAD,active_validation=False)
    mylogger(f'response: {response}')
    return {"executionId": response.get('executionId')}

def ReserveLocalPool(PAYLOAD, AuthKey):
    URL = f'https://{HOST}/api/v2/ippool/group'
    HEADERS = {
            'content-type': "application/json",
            'x-auth-token': AuthKey
        }
    mylogger(f'About to create pool with payload: {PAYLOAD}')
    response = requests.post(url=URL, headers=HEADERS, verify=False, data=json.dumps(PAYLOAD))
    mylogger(f'response.text: {response.text}')
    return {"taskId": response.json().get('response').get('taskId')}

def validateJSON(jsonData):
    try:
        return json.loads(jsonData)
    except ValueError as err:
        return False
    #return True

def CreateDeviceCredentialsSDK(PAYLOAD,DNAC_SDK):
    mylogger(f'About to create device credentials with payload: {PAYLOAD}')
    response = DNAC_SDK.network_settings.create_device_credentials(payload=PAYLOAD,active_validation=False)
    mylogger(f'response: {response}')

def ImportCredentials(xlSheet, DNAC_SDK):
    Responses = []
    for row in range(2, xlSheet.max_row+1):
        column = 1
        for cell in xlSheet[row]:
            if column == 2:
                SiteHierarchy = cell.value
            elif column == 3:
                SiteType = cell.value
            elif column == 4 and SiteType.lower() == 'building':
                SiteAddress = cell.value
            column += 1
        mylogger(f'{lineno()} SiteHierarchy value: {SiteHierarchy} Row: {row}')
        SplitSite = SiteHierarchy.split("/")
        if isinstance(SplitSite,list):
            SiteName = SplitSite[len(SplitSite) - 1]
            SplitSite.pop()
            ParentName = "Global/" + '/'.join(SplitSite)
        else:
            ParentName = "Global"
            SiteName = SiteHierarchy

def ImportSites(xlSheet, DNAC_SDK):
    SiteType = ''
    ParentName = ''
    Responses = []
    for row in range(2, xlSheet.max_row+1):
        column = 1
        for cell in xlSheet[row]:
            if column == 2:
                SiteHierarchy = cell.value
            elif column == 3:
                SiteType = cell.value
            elif column == 4 and SiteType.lower() == 'building':
                SiteAddress = cell.value
            elif column == 5 and SiteType.lower() == 'floor':
                FloorRfModel = cell.value
            elif column == 6 and SiteType.lower() == 'floor':
                FloorWidth = cell.value
            elif column == 7 and SiteType.lower() == 'floor':
                FloorLength = cell.value
            elif column == 8 and SiteType.lower() == 'floor':
                FloorHeight = cell.value
            column += 1
        mylogger(f'{lineno()} SiteHierarchy value: {SiteHierarchy} Row: {row}')
        SplitSite = SiteHierarchy.split("/")
        if isinstance(SplitSite,list):
            SiteName = SplitSite[len(SplitSite) - 1]
            SplitSite.pop()
            ParentName = "Global/" + '/'.join(SplitSite)
        else:
            ParentName = "Global"
            SiteName = SiteHierarchy
        
        if SiteType.lower() == 'building':
            PAYLOAD = {
                "building": {
                    "name": SiteName,
                    "parentName": ParentName,
                    "address": SiteAddress
                }
            }
            #populate responses to check their status           
            Responses.append(CreateSiteSDK(DNAC_SDK, PAYLOAD, "building"))
            Responses[len(Responses) - 1].update({'SiteName': SiteName})
        elif SiteType.lower() == 'floor':
            PAYLOAD = {
                "floor": {
                    "name": SiteName,
                    "parentName": ParentName,
                    "rfModel": FloorRfModel,
                    "width": FloorWidth,
                    "length": FloorLength,
                    "height": FloorHeight
                }
            }
            #populate responses to check their status           
            Responses.append(CreateSiteSDK(DNAC_SDK, PAYLOAD, "floor"))
            Responses[len(Responses) - 1].update({'SiteName': SiteName})
        else:
            PAYLOAD = {
                "area": {
                    "name": SiteName,
                    "parentName": ParentName
                }
            }
            Responses.append(CreateSiteSDK(DNAC_SDK, PAYLOAD, "area"))
            Responses[len(Responses) - 1].update({'SiteName': SiteName})
        
        sleep(1)
    checkResponses(Responses, DNAC_SDK)

def ImportSitesWithYield(xlSheet, DNAC_SDK):
    mylogger(f'{lineno()} Beginning to run ImportSitesWithYield')
    SiteType = ''
    ParentName = ''
    Responses = []
    for row in range(2, xlSheet.max_row+1):
        column = 1
        for cell in xlSheet[row]:
            if column == 2:
                SiteHierarchy = cell.value
            elif column == 3:
                SiteType = cell.value
            elif column == 4 and SiteType.lower() == 'building':
                SiteAddress = cell.value
            elif column == 5 and SiteType.lower() == 'floor':
                FloorRfModel = cell.value
            elif column == 6 and SiteType.lower() == 'floor':
                FloorWidth = cell.value
            elif column == 7 and SiteType.lower() == 'floor':
                FloorLength = cell.value
            elif column == 8 and SiteType.lower() == 'floor':
                FloorHeight = cell.value
            column += 1
        mylogger(f'{lineno()} SiteHierarchy value: {SiteHierarchy} Row: {row}')
        SplitSite = SiteHierarchy.split("/")
        if isinstance(SplitSite,list):
            SiteName = SplitSite[len(SplitSite) - 1]
            SplitSite.pop()
            ParentName = "Global/" + '/'.join(SplitSite)
        else:
            ParentName = "Global"
            SiteName = SiteHierarchy
        
        if SiteType.lower() == 'building':
            PAYLOAD = {
                "building": {
                    "name": SiteName,
                    "parentName": ParentName,
                    "address": SiteAddress
                }
            }
            #populate responses to check their status           
            Responses.append(CreateSiteSDK(DNAC_SDK, PAYLOAD, "building"))
            Responses[len(Responses) - 1].update({'SiteName': SiteName})
        elif SiteType.lower() == 'floor':
            PAYLOAD = {
                "floor": {
                    "name": SiteName,
                    "parentName": ParentName,
                    "rfModel": FloorRfModel,
                    "width": FloorWidth,
                    "length": FloorLength,
                    "height": FloorHeight
                }
            }
            #populate responses to check their status           
            Responses.append(CreateSiteSDK(DNAC_SDK, PAYLOAD, "floor"))
            Responses[len(Responses) - 1].update({'SiteName': SiteName})
        else:
            PAYLOAD = {
                "area": {
                    "name": SiteName,
                    "parentName": ParentName
                }
            }
            Responses.append(CreateSiteSDK(DNAC_SDK, PAYLOAD, "area"))
            Responses[len(Responses) - 1].update({'SiteName': SiteName})
        
        yield f'Created site: {SiteName}'
        sleep(1)
    checkResponses(Responses, DNAC_SDK)
    return "done"

def DeleteSite(siteId, DNAC_SDK):
    myResponse = DNAC_SDK.sites.delete_site(siteId)
    checkResponses([myResponse], DNAC_SDK)

def DeleteSitesFromSheet(xlSheet, DNAC_SDK, iterate = False):
    SiteType = ''
    ParentName = ''
    Responses = []
    SiteNames = []
    for row in range(2, xlSheet.max_row+1):
        column = 1
        for cell in xlSheet[row]:
            if column == 2:
                SiteHierarchy = cell.value
            column += 1
        mylogger(f'{lineno()} SiteHierarchy value: {SiteHierarchy} Row: {row}')
        FullSiteName = "Global/" + SiteHierarchy
        SiteNames.append(FullSiteName)
    
    AllSites = getSitesSDK(DNAC_SDK)
    x = 1
    while x <= len(SiteNames):
        try:
            mySiteId = AllSites[SiteNames[len(SiteNames) - x]]
            mylogger(f'{lineno()} Site: {SiteNames[len(SiteNames) - x]} being deleted')
            DeleteSite(mySiteId, DNAC_SDK)
            if iterate:
                yield SiteNames[len(SiteNames) - x]
            mylogger(f'Deleted site: {SiteNames[len(SiteNames) - x]}')
        except KeyError:
            mylogger(f'{lineno()} Site: {SiteNames[len(SiteNames) - x]} not found for deletion')
            pass
        x = x + 1

def checkResponses(Responses, DNAC_SDK):
    for response in Responses:
        mylogger(f'{lineno()} Getting status of {response}')
        if response.get('executionId'):
            getExecutionStatus(response['executionId'],DNAC_SDK.access_token)
        elif response.get('taskId'):
            getTaskStatus(response['taskId'],DNAC_SDK.access_token)

def ImportPools(xlSheet, DNAC_SDK):
    PoolType = ''
    ParentName = ''
    mySites = getSitesSDK(DNAC_SDK)
    Responses = []
    for row in range(2, xlSheet.max_row+1):
        column = 1
        #GotSites = False
        PoolProperties = {'PoolName': None, 'PoolSubnet': None, 'PoolSite': None, 'GlobalPoolName': None, 'PoolDhcpServers': None, 'PoolDnsServers': None, 'PoolGateway': None}
        withDHCP = False
        withDNS = False
        withGateway = False
        for cell in xlSheet[row]:
            if column == 1:
                PoolProperties['PoolName'] = cell.value
            elif column == 2:
                PoolProperties['PoolSubnet'] = cell.value
            elif column == 3:
                PoolProperties['PoolSite'] = cell.value
            elif column == 4 and PoolProperties['PoolSite'].lower() != 'global':
                PoolProperties['GlobalPoolName'] = cell.value
            elif column == 5 and PoolProperties['PoolSite'].lower() != 'global':
                if cell.value:
                    PoolProperties['PoolDhcpServers'] = [cell.value]
                    withDHCP = True
            elif column == 6 and PoolProperties['PoolSite'].lower() != 'global':
                if cell.value:
                    PoolProperties['PoolDnsServers'] = [cell.value]
                    withDNS = True
            elif column == 7 and PoolProperties['PoolSite'].lower() != 'global':
                if cell.value:
                    PoolProperties['PoolGateway'] = [cell.value]
                    withGateway = True
            column += 1
        #Remove Nothing items for logging
        PoolPropertiesValues = (item for item in PoolProperties.values() if item)
        for item in PoolPropertiesValues:
            mylogger(f'{lineno()} The item: {item} Row: {row}')
        #mylogger(f'{lineno()} Current Pool Properties: {PoolPropertiesValues} Row: {row}')
            
        if PoolProperties['PoolSite'].lower() == 'global':
            mypool = {
                "settings": {
                    "ippool": [
                        {                
                            "ipPoolName": PoolProperties['PoolName'],
                            "type": "Generic",
                            "ipPoolCidr": PoolProperties['PoolSubnet'],
                            "IpAddressSpace": "IPv4"
                        }
                    ]
                }
            }     
            #populate responses to check their status           
            Responses.append(CreateGlobalPoolSDK(mypool,DNAC_SDK))
            Responses[len(Responses) - 1].update({'Cidr': PoolProperties['PoolSubnet']})
        else:
            print(f'{lineno()} Global Pool ID: {mySites["Global/" + PoolProperties["PoolSite"]]}')
            GlobalPoolId = getGlobalPoolId(PoolProperties['GlobalPoolName'],DNAC_SDK.access_token)
            if withGateway and withDNS and withDHCP:
                PAYLOAD = {
                    "groupName":PoolProperties['PoolName'],
                    "groupOwner":"DNAC",
                    "type":"generic",
                    "siteId":mySites["Global/" + PoolProperties['PoolSite']],
                        "ipPools":
                        [
                            {
                                "parentUuid": GlobalPoolId,
                                "dhcpServerIps": PoolProperties['PoolDhcpServers'],
                                "dnsServerIps": PoolProperties['PoolDnsServers'],
                                "ipPoolOwner": "DNAC",
                                "shared": True,
                                "gateways": PoolProperties['PoolGateway'],
                                "ipPoolCidr": PoolProperties['PoolSubnet']
                            }
                        ]
                    }
            elif withGateway and withDHCP:
                PAYLOAD = {
                    "groupName":PoolProperties['PoolName'],
                    "groupOwner":"DNAC",
                    "type":"generic",
                    "siteId":mySites["Global/" + PoolProperties['PoolSite']],
                        "ipPools":
                        [
                            {
                                "parentUuid": GlobalPoolId,
                                "dhcpServerIps": PoolProperties['PoolDhcpServers'],
                                "ipPoolOwner": "DNAC",
                                "shared": True,
                                "gateways": PoolProperties['PoolGateway'],
                                "ipPoolCidr": PoolProperties['PoolSubnet']
                            }
                        ]
                    }
            elif withGateway and withDNS:
                PAYLOAD = {
                    "groupName":PoolProperties['PoolName'],
                    "groupOwner":"DNAC",
                    "type":"generic",
                    "siteId":mySites["Global/" + PoolProperties['PoolSite']],
                        "ipPools":
                        [
                            {
                                "parentUuid": GlobalPoolId,
                                "dnsServerIps": PoolProperties['PoolDnsServers'],
                                "ipPoolOwner": "DNAC",
                                "shared": True,
                                "gateways": PoolProperties['PoolGateway'],
                                "ipPoolCidr": PoolProperties['PoolSubnet']
                            }
                        ]
                    }
            elif withDNS and withDHCP:
                PAYLOAD = {
                    "groupName":PoolProperties['PoolName'],
                    "groupOwner":"DNAC",
                    "type":"generic",
                    "siteId":mySites["Global/" + PoolProperties['PoolSite']],
                        "ipPools":
                        [
                            {
                                "parentUuid": GlobalPoolId,
                                "dhcpServerIps": PoolProperties['PoolDhcpServers'],
                                "dnsServerIps": PoolProperties['PoolDnsServers'],
                                "ipPoolOwner": "DNAC",
                                "shared": True,
                                "ipPoolCidr": PoolProperties['PoolSubnet']
                            }
                        ]
                    }
            elif withGateway:
                PAYLOAD = {
                    "groupName":PoolProperties['PoolName'],
                    "groupOwner":"DNAC",
                    "type":"generic",
                    "siteId":mySites["Global/" + PoolProperties['PoolSite']],
                        "ipPools":
                        [
                            {
                                "parentUuid": GlobalPoolId,
                                "ipPoolOwner": "DNAC",
                                "shared": True,
                                "gateways": PoolProperties['PoolGateway'],
                                "ipPoolCidr": PoolProperties['PoolSubnet']
                            }
                        ]
                    }
            elif withDNS:
                PAYLOAD = {
                    "groupName":PoolProperties['PoolName'],
                    "groupOwner":"DNAC",
                    "type":"generic",
                    "siteId":mySites["Global/" + PoolProperties['PoolSite']],
                        "ipPools":
                        [
                            {
                                "parentUuid": GlobalPoolId,
                                "dnsServerIps": PoolProperties['PoolDnsServers'],
                                "ipPoolOwner": "DNAC",
                                "shared": True,
                                "ipPoolCidr": PoolProperties['PoolSubnet']
                            }
                        ]
                    }
            elif withDHCP:
                PAYLOAD = {
                    "groupName":PoolProperties['PoolName'],
                    "groupOwner":"DNAC",
                    "type":"generic",
                    "siteId":mySites["Global/" + PoolProperties['PoolSite']],
                        "ipPools":
                        [
                            {
                                "parentUuid": GlobalPoolId,
                                "dhcpServerIps": PoolProperties['PoolDhcpServers'],
                                "ipPoolOwner": "DNAC",
                                "shared": True,
                                "ipPoolCidr": PoolProperties['PoolSubnet']
                            }
                        ]
                    }
            else:
                PAYLOAD = {
                    "groupName":PoolProperties['PoolName'],
                    "groupOwner":"DNAC",
                    "type":"generic",
                    "siteId":mySites["Global/" + PoolProperties['PoolSite']],
                        "ipPools":
                        [
                            {
                                "parentUuid": GlobalPoolId,
                                "ipPoolOwner": "DNAC",
                                "shared": True,
                                "ipPoolCidr": PoolProperties['PoolSubnet']
                            }
                        ]
                    }
            Responses.append(ReserveLocalPool(PAYLOAD, DNAC_SDK.access_token))
            Responses[len(Responses) - 1].update({'Cidr': PoolProperties['PoolSubnet']})
        sleep(1)
    checkResponses(Responses, DNAC_SDK)
        
def getTaskStatus(taskId, AuthKey):
    URL = f'https://{HOST}/api/v1/task/{taskId}'
    HEADERS = {
            'content-type': "application/json",
            'x-auth-token': AuthKey
        }
    mylogger(f'{lineno()} About to retrieve task status for task: {taskId}')
    response = requests.get(url=URL, headers=HEADERS, verify=False)
    mylogger(f'{lineno()} response.text: {response.text}')

def getExecutionStatus(taskId, AuthKey):
    URL = f'https://{HOST}/dna/platform/management/business-api/v1/execution-status/{taskId}'

    HEADERS = {
            'content-type': "application/json",
            'x-auth-token': AuthKey
        }
    mylogger(f'{lineno()} About to retrieve execution status for task: {taskId}')
    response = requests.get(url=URL, headers=HEADERS, verify=False)
    mylogger(f'{lineno()} response.text: {response.text}')
    if response.json()['status'] == "IN_PROGRESS":
        mylogger(f'{lineno()} Response in progress. Sleep 1 second and try again')
        print(f'Response in progress. Will retry in 1 second')
        sleep(2)
        getExecutionStatus(taskId, AuthKey)

def createglobalpool(DNAC_SDK):
    isSuccess = False
    while not isSuccess:
        myPoolName = input("Enter IP Pool Name: ")
        myPoolCider = input("Enter IP Pool CIDR (###.###.###.###/##): ")
        try:
            if myPoolName.isalnum():
                mylogger("good pool name")
                isSuccess = True
            else:
                pprint.pprint("Please ensure only Numbers, Letters, or spaces are used")
                continue
            if ipaddress.ip_network(myPoolCider,strict=True):
                mylogger("good ip network")
                isSuccess = True
        except Exception as myE:
            mylogger(myE)
            pprint.pprint(traceback.print_exc())
            pprint.pprint("Please enter a valid IP prefix")
        mypool = {
            "settings": {
                "ippool": [
                    {                
                        "ipPoolName": myPoolName,
                        "type": "Generic",
                        "ipPoolCidr": myPoolCider,
                        "IpAddressSpace": "IPv4"
                    }
                ],
                "IpAddressSpace": "IPv4"
            }
        }                
        #CreateGlobalPool(mypool,TOKEN)
        CreateGlobalPoolSDK(mypool,DNAC_SDK)

def createdevicecredentials(DNAC_SDK):
    isSuccess = False
    fromFile = False
    while not isSuccess:
        print("Enter/Paste your JSON content. Enter 'from file' to read from file. Ctrl-D or Ctrl-Z ( windows ) to end.")
        contents = []
        while True:
            try:
                line = input()
                if 'from file' in line.lower():
                    FileName = input('Please enter file name of properly formatted json file:')
                    with open(FileName, 'r') as f:
                        PAYLOAD = json.load(f)
                    isSuccess = True
                    fromFile = True
                    break
            except EOFError:
                break
            contents.append(line)
        if not fromFile:
            myJsonInput = '\n'.join(contents)
            PAYLOAD = validateJSON(myJsonInput)
            if PAYLOAD:
                isSuccess = True
            else:
                print("Try again. ")
    CreateDeviceCredentialsSDK(PAYLOAD,DNAC_SDK)
    myResponse = input("Assign credentials to a site? (Yes/No) <No>: ")
    if "yes" in myResponse.lower():
        sites = {}
        sites = getSitesSDK(DNAC_SDK)
        print('Enter site to assign these credentials to.')
        print('\n'.join(sites.keys()))
        input(": ")

def main(argv):
    try:
        # set default values for command line arguments
        initlogging(argv)
        global username
        global password
        global HOST
        myAction = CheckUsage(argv)
        while not myAction:
            newAction = input('Please choose option: ')
            myArgv = [argv, newAction]
            myAction = CheckUsage(myArgv)
            
        csv_file1 = None
        myTokenSuccess = False
        while not myTokenSuccess:
            try:
                import login
                if hasattr(login, 'username'):
                    username = login.username
                if hasattr(login, 'password'):
                    password = login.password
                if hasattr(login, 'HOST'):
                    HOST = login.HOST
                if username == '':
                    username = input('Enter your DNAC username: ')
                if password == '':
                    password = getpass('Enter your DNAC password: ')
                if HOST == '':
                    HOST = input("Please enter the DNAC's IP address: ")
            except ImportError:
                username = input('Enter your DNAC username: ')
                password = getpass('Enter your DNAC password: ')
                HOST = input("Please enter the DNAC's IP address: ")
            #using DNAC API to retrieve token
            #TOKEN = getAuthToken(username, password)    
            i = 0
            try:
                dnac_api = DNACenterAPI(username=username, password=password, base_url=f'https://{HOST}', verify=False)
                myTokenSuccess = True
            except ApiError as myErr:
                print(myErr)
            TOKEN = dnac_api.access_token
        if myArgv[1].lower() == 'importpoolsfromcsv':
            IpPoolsList = []
            with open(argv[2], 'r') as mycsv_file:
                myfieldnames = ['Pool Name', 'Subnet', 'Site Name', 'Global Pool Name', 'DHCP Server IPs', 'DNS Server IPs', 'Default Gateway']
                myreader = csv.DictReader(mycsv_file, myfieldnames)
                for row in myreader:
                    if i == 0:
                        mylogger('Skipping row 1')
                        i = i + 1
                    else:
                        IpPoolsList.append(row)
                        #mylogger(row)
        
            SiteName = ''
            GlobalPoolName = ''
            SiteId = ''
            GlobalPoolId = ''
            for pool in IpPoolsList:
                mylogger(f'Begin processing site: {pool["Site Name"]}')
                if not SiteName == pool['Site Name'] or SiteId == '':
                    SiteName = pool['Site Name']
                    SiteId = getSiteId(pool['Site Name'], TOKEN)
                if not GlobalPoolName == pool['Global Pool Name'] or GlobalPoolId == '':
                    GlobalPoolName = pool['Global Pool Name']
                    GlobalPoolId = getGlobalPoolId(pool['Global Pool Name'], TOKEN)
                PAYLOAD = {
                    "groupName":pool["Pool Name"],
                    "groupOwner":"DNAC",
                    "type":"generic",
                    "siteId":SiteId,
                        "ipPools":
                        [
                            {
                                "parentUuid":GlobalPoolId,
                                "dhcpServerIps":[pool["DHCP Server IPs"]],
                                "dnsServerIps":[pool["DNS Server IPs"]],
                                "ipPoolOwner":"DNAC",
                                "shared":True,
                                "gateways":[pool["Default Gateway"]],
                                "ipPoolCidr":pool["Subnet"]
                            }
                        ]
                    }
                ReserveLocalPool(PAYLOAD, TOKEN)
        elif myArgv[1].lower() == 'createglobalpool':
            createglobalpool(dnac_api)
        
        elif myArgv[1].lower() == 'createdevicecredentials':
            createdevicecredentials(dnac_api)
        
        elif myArgv[1].lower() == 'createnetworksettings':
            pass

        elif myArgv[1].lower() == 'importfromexcel':
            filename = input('Excel file name (xlsx): ')
            import openpyxl as xl
            mywb = xl.load_workbook(filename)

            for sheet in mywb.sheetnames:
                if sheet == 'Sites':
                    ImportSites(mywb[sheet], dnac_api)
                #elif sheet == 'Pools':
                #    ImportPools(mywb[sheet], dnac_api)
                #elif sheet == 'Credentials':
                #    ImportCredentials(mywb[sheet], dnac_api)

        elif myArgv[1].lower() == 'deletefromexcel':
            filename = input('Excel file name (xlsx): ')
            import openpyxl as xl
            mywb = xl.load_workbook(filename)

            for sheet in mywb.sheetnames:
                if sheet == 'Sites':
                    DeleteSitesFromSheet(mywb[sheet], dnac_api)
                


        print(f'Base file name of logger is: {logfile}')
    except KeyboardInterrupt:
        print(f'Exception: base file name of logger is: {logfile}')



    

    


if __name__ == '__main__':
    main(sys.argv)
