from flask import Flask, json, request, Response, render_template, stream_with_context
import json, sys, requests, re
import os
import datetime
from pytz import timezone
import dateutil.parser
from pprint import pprint
import re
from dnacentersdk import DNACenterAPI, ApiError
import logging
from logging.handlers import RotatingFileHandler
import inspect
import DNAC_Utility_v03
import time, math
import random, string
import cveLogger

myTokenSuccess = False
dnac_api = ''
app = Flask(__name__)
app.secret_key = ''.join(random.SystemRandom().choice(string.ascii_letters + string.digits) for _ in range(18))

@app.route('/')
def index():
    return render_template("login.html")


@app.route("/", methods=["POST"])
def login_dnac():

    # Webhook Receiver
    webhook_data = request.json
    #pprint(webhook_data)
    #webhook_data = json.dumps(webhook_data)


@app.route("/index", methods=['POST'])
def renderIndex():
    global myTokenSuccess
    global dnac_api
    code = request.form.get("code")
    cveLogger.mylogger(f"this is the code: {code}")

     #using DNAC API to retrieve token
    #TOKEN = getAuthToken(username, password)  
    username = request.form.get("username") 
    password = request.form.get("password")
    HOST = request.form.get("IpAddress")
    DNAC_Utility_v03.HOST = HOST
    i = 0
    try:
        cveLogger.mylogger(f'about to try DNAC with address: {HOST}')
        dnac_api = DNACenterAPI(username=username, password=password, base_url=f'https://{HOST}', verify=False)
        myTokenSuccess = True
    except ApiError as myErr:
        cveLogger.mylogger(f'API Error: {myErr}')
        return render_template("login.html")
    TOKEN = dnac_api.access_token

    cveLogger.mylogger(request.data)
    cveLogger.mylogger(f'{cveLogger.lineno()} {request.get_data()}')
    cveLogger.mylogger(f'{cveLogger.lineno()} {request.get_json()}')
    for myvalue in request.form.values():
        cveLogger.mylogger(f'{cveLogger.lineno()} {myvalue}')
    #webhook_data = json.dumps(webhook_data)

    return render_template("importSites.html.jinja", orgs = "organizations")

@app.route("/importSites")
def importSites():
    if dnac_api == '':
        return render_template("login.html")
    try:
        return render_template("importSites.html.jinja", orgs = "organizations")
    except Exception as myE:
        cveLogger.mylogger(f'{cveLogger.lineno()} myE Error: {myE}')
        return render_template("importSites.html.jinja", orgs = "organizations")

@app.route("/importStatus", methods=['POST', 'GET'])
def import_post():
    cveLogger.mylogger(f'{cveLogger.lineno()} In import_post(). About to create generate method')  
    @stream_with_context
    def generate():
        global dnac_api
        import openpyxl as xl
        filename = 'uploads/myimportsites.xlsx'
        mywb = xl.load_workbook(filename)

        for sheet in mywb.sheetnames:
            if sheet == 'Sites':
                cveLogger.mylogger(f'{cveLogger.lineno()} About to run DNAC_Utility ImportSitesWithYield. Sheet: {sheet}')
                yield ' '*1024 + '\n'
                for mysite in DNAC_Utility_v03.ImportSitesWithYield(mywb[sheet], dnac_api):
                    cveLogger.mylogger(f'{cveLogger.lineno()} Yielding: {mysite}') 
                    yield mysite + '\n'
                    cveLogger.mylogger(f'{cveLogger.lineno()} After yield')
        if os.path.exists(filename):
            os.remove(filename)
            cveLogger.mylogger(f'{cveLogger.lineno()} file "{filename}" was deleted')
        else:
            cveLogger.mylogger(f'{cveLogger.lineno()} file "{filename}" does not exist')

    if request.method == "POST":
        cveLogger.mylogger(f'{cveLogger.lineno()} request method is post')
        cveLogger.mylogger(f'{cveLogger.lineno()} Request content type: {request.content_type}')

        filename = 'uploads/myimportsites.xlsx'
        for myfile in request.files:
            cveLogger.mylogger(f'{cveLogger.lineno()} Has a file: {myfile}')
        try: 
            request.files['file-select'].save(filename)
            cveLogger.mylogger(f'{cveLogger.lineno()} Successfully saved: {filename}')            
        except Exception as myE:
            cveLogger.mylogger(f'{cveLogger.lineno()} myE Error: {myE}')
            return "error saving file"
        return render_template("importStatus.html.jinja", orgs = "organizations")
    else:                
        cveLogger.mylogger(f'{cveLogger.lineno()} request method is not post')  
        return app.response_class(generate(), mimetype="text/plain")

@app.route("/deleteSites")
def deleteSites():
    if dnac_api == '':
        return render_template("login.html")
    return render_template("deleteSites.html.jinja")

@app.route("/deleteStatus", methods=['POST', 'GET'])
def delete_post():
    
    def generate():
        global dnac_api
        filename = 'uploads/mydelfile.xlsx'
        import openpyxl as xl
        mywb = xl.load_workbook(filename)

        for sheet in mywb.sheetnames:
            if sheet == 'Sites':
                cveLogger.mylogger(f'{cveLogger.lineno()} About to run DNAC_Utility DeleteSitesFromSheet. Sheet: {sheet}')
                yield ' '*1024 + '\n'
                for mysite in DNAC_Utility_v03.DeleteSitesFromSheet(mywb[sheet], dnac_api, iterate = True):
                    cveLogger.mylogger(f'{cveLogger.lineno()} Yielding: {mysite}') 
                    yield mysite + '\n'
                    cveLogger.mylogger(f'{cveLogger.lineno()} After yield')
        if os.path.exists(filename):
            os.remove(filename)
            cveLogger.mylogger(f'{cveLogger.lineno()} file "{filename}" was deleted')
        else:
            cveLogger.mylogger(f'{cveLogger.lineno()} file "{filename}" does not exist')
                    
    if request.method == "POST":
        cveLogger.mylogger(f'{cveLogger.lineno()} request method is post')
        cveLogger.mylogger(f'{cveLogger.lineno()} Request content type: {request.content_type}')
        filename = 'uploads/mydelfile.xlsx'
        for myfile in request.files:
            cveLogger.mylogger(f'{cveLogger.lineno()} Has a file: {myfile}')
        try: 
            request.files['file-select'].save(filename)
            
        except Exception as myE:
            cveLogger.mylogger(f'{cveLogger.lineno()} API Error: {myE}')
            return render_template("deleteSites.html.jinja")
        return render_template("deleteStatus.html.jinja", orgs = "organizations")
    else:                
        cveLogger.mylogger(f'{cveLogger.lineno()} request method is not post')  
        return app.response_class(generate(), mimetype="text/plain")

    return sys.argv[0]

@app.route("/jobStatus", methods=['GET'])
def getJobStatus():
    return "not implemented"
    
#this next block is for testing
# 
@app.route('/myContent') # render the content a url differnt from index
def content():
    def inner():
        # simulate a long process to watch
        for i in range(100):
            j = math.sqrt(i)
            time.sleep(1)
            # this value should be inserted into an HTML template
            yield str(i) + '<br/>\n'
    return Response(inner(), mimetype='text/html')

@app.route('/testIndex')
def testIndex():
    return render_template('testIndex.html') # render a template at the index. The content will be embedded in this template

@app.route("/testIndex2")
def testIndex2():
    return render_template("testIndex2.html.jinja")

@app.route("/stream")
def stream():
    def generate():
        #for i in range(100):
        #   cveLogger.mylogger(f'{cveLogger.lineno()} generating for i value: {i}')
        #   yield "{}\n".format(math.sqrt(i))
        #   cveLogger.mylogger(f'{cveLogger.lineno()} about to sleep')
        #   time.sleep(1)
        cveLogger.mylogger(f'{cveLogger.lineno()} going to call outside generator')
        for squarert in generatehere():
            yield squarert
        cveLogger.mylogger(f'{cveLogger.lineno()} past outside generator')

    cveLogger.mylogger(f'{cveLogger.lineno()} about to run generator')
    return app.response_class(generate(), mimetype="text/plain")

#end testing block    

@app.route("/settings")
def settings():
    return render_template("settings.html")

def generatehere():
    for i in range(100):
        cveLogger.mylogger(f'{cveLogger.lineno()} generating for i value: {i}')
        yield "{}\n".format(math.sqrt(i))
        cveLogger.mylogger(f'{cveLogger.lineno()} about to sleep')
        time.sleep(1)

if __name__ == "__main__":
    cveLogger.initlogging(sys.argv)
    cveLogger.mylogger(f'{cveLogger.lineno()} About to begin web server on port 5001')
    app.run(host="0.0.0.0", port=5001, debug=True)